from tkinter import *
from openpyxl import *
from tkinter import messagebox

if __name__ == '__main__':
    root = Tk()
    root.geometry("800x600")
    root.maxsize(750, 600)
    root.minsize(750, 600)
    root.config(bg="#676c73")
    wb = Workbook()
    sheet = wb.active
    root.title("LIBRARY MANAGEMENT SYSTEM")


    firstName_var = StringVar()
    middleName_var = StringVar()
    lastName_var = StringVar()
    contactNo1_var = StringVar()
    contactNo2_var = StringVar()
    email_var = StringVar()
    village_var = StringVar()
    post_var = StringVar()
    polishStation_var = StringVar()
    pinCode_var = StringVar()
    ######################################
    bookName_var = StringVar()
    bookAuthorName_var = StringVar()
    bookId_var = StringVar()
    bookIssueDate_var = StringVar()
    bookPrice_var = StringVar()
    bookSubmitDate_var = StringVar()
    bookLateSubmitDate_var = StringVar()
    booklateSubmmitFine_var = StringVar()
    totalPricepad_var = StringVar()

    font1 = ('calibre', 10, 'bold')
    impformationList = ["list of data", "First Name", "Middle Name", "Last Name", "Contact No-1", "Contact No-2",
                        "Email",
                        "Village", "Post", "Pin Code", "Book Name", "Book Author Name", "Book Id No.",
                        "Book Issue Date",
                        "Book Price", "Book submit Date", "Book late Submit Date", "Late Submit fine", "Total Price"]
    for il in range(1, len(impformationList)):
        sheet.cell(row=1, column=il).value = impformationList[il]
        print(il, " " + impformationList[il])

    i = 1

    Title_Lable=Label(root, text="LIBRARY MANAGEMENT SYSTEM", font="calibre 17 bold",bg="#1adbae").place(x=150, y=20)

    def submit():
        """"This is a entry box for in exel"""
        global i
        correctfill = True
        i += 1
        sheet.cell(row=i, column=1).value = firstName_var.get()
        if len(firstName_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the First Name")

        sheet.cell(row=i, column=2).value = middleName_var.get()

        sheet.cell(row=i, column=3).value = lastName_var.get()
        if len(lastName_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Last Name")

        sheet.cell(row=i, column=4).value = contactNo1_var.get()
        if len(contactNo1_var.get()) != 10:
            correctfill = False
            messagebox.showerror("ERROR", "Invalid contact Number-1")

        sheet.cell(row=i, column=5).value = contactNo2_var.get()

        sheet.cell(row=i, column=6).value = email_var.get()
        if len(email_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Email")

        sheet.cell(row=i, column=7).value = village_var.get()
        if len(village_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Village Name")

        sheet.cell(row=i, column=8).value = post_var.get()
        if len(post_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Post Name")

        sheet.cell(row=i, column=9).value = pinCode_var.get()
        if len(pinCode_var.get()) != 7:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the valid Pine Code")

        sheet.cell(row=i, column=10).value = bookName_var.get()
        if len(bookName_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Name")

        sheet.cell(row=i, column=11).value = bookAuthorName_var.get()
        if len(bookAuthorName_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Author Name")

        sheet.cell(row=i, column=12).value = bookId_var.get()
        if len(bookId_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Id Number")

        sheet.cell(row=i, column=13).value = bookIssueDate_var.get()
        if len(bookIssueDate_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Issue Date")

        sheet.cell(row=i, column=14).value = bookPrice_var.get()
        if len(bookPrice_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Price")

        sheet.cell(row=i, column=15).value = bookSubmitDate_var.get()
        if len(bookSubmitDate_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book Submit Date")

        sheet.cell(row=i, column=16).value = bookLateSubmitDate_var.get()
        if len(bookLateSubmitDate_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book late Submit Date")

        sheet.cell(row=i, column=17).value = booklateSubmmitFine_var.get()
        if len(booklateSubmmitFine_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Book late Submit Fine")

        sheet.cell(row=i, column=18).value = totalPricepad_var.get()
        if len(totalPricepad_var.get()) == 0:
            correctfill = False
            messagebox.showerror("ERROR", "Please Enter the Total price Pad ")
        if correctfill:
            wb.save("C:\\Users\\vishn\\OneDrive\\Desktop\\raushtttt11.xlsx")


    # name_label = Label(root, text='Username', font=font1).grid(row=0, column=0)
    # name_entry = Entry(root, textvariable=name_var, font=font1).grid(row=0, column=1)

    firstName_lable = Label(root, text="First Name", font=font1).place(x=50, y=100)
    firstName_Entry = Entry(root, textvariable=firstName_var, font=font1).place(x=150, y=100)

    middleName_lable = Label(root, text="Middle Name", font=font1).place(x=50, y=150)
    middleName_Entry = Entry(root, textvariable=middleName_var, font=font1).place(x=150, y=150)

    lastName_lable = Label(root, text="Last Name", font=font1).place(x=50, y=200)
    lastName_Entry = Entry(root, textvariable=lastName_var, font=font1).place(x=150, y=200)

    contactNo1_lable = Label(root, text="Contact No-1", font=font1).place(x=50, y=250)
    contactNo1_Entry = Entry(root, textvariable=contactNo1_var, font=font1).place(x=150, y=250)

    contactNo2_lable = Label(root, text="Contact No-2", font=font1).place(x=50, y=300)
    contactNo2_Entry = Entry(root, textvariable=contactNo2_var, font=font1).place(x=150, y=300)

    email_lable = Label(root, text="Email", font=font1).place(x=50, y=350)
    email_Entry = Entry(root, textvariable=email_var, font=font1).place(x=150, y=350)

    village_lable = Label(root, text="Village", font=font1).place(x=50, y=400)
    village_Entry = Entry(root, textvariable=village_var, font=font1).place(x=150, y=400)

    post_lable = Label(root, text="Post", font=font1).place(x=50, y=450)
    post_Entry = Entry(root, textvariable=post_var, font=font1).place(x=150, y=450)

    pinCode_lable = Label(root, text="Pin Code", font=font1).place(x=50, y=500)
    pinCode_Entry = Entry(root, textvariable=pinCode_var, font=font1).place(x=150, y=500)

    ####################################################################

    bookName_lable = Label(root, text="Book Name", font=font1).place(x=400, y=100)
    bookName_Entry = Entry(root, textvariable=bookName_var, font=font1).place(x=550, y=100)

    bookAuthorName_lable = Label(root, text="Book Author Name", font=font1).place(x=400, y=150)
    bookAuthorName_Entry = Entry(root, textvariable=bookAuthorName_var, font=font1).place(x=550, y=150)

    bookId_lable = Label(root, text="Book Id No.", font=font1).place(x=400, y=200)
    bookId_Entry = Entry(root, textvariable=bookId_var, font=font1).place(x=550, y=200)

    bookIssueDate_lable = Label(root, text="Book Issue Date", font=font1).place(x=400, y=250)
    bookIssueDate_Entry = Entry(root, textvariable=bookIssueDate_var, font=font1).place(x=550, y=250)

    bookPrice_lable = Label(root, text="Book Price", font=font1).place(x=400, y=300)
    bookPrice_Entry = Entry(root, textvariable=bookPrice_var, font=font1).place(x=550, y=300)

    bookSubmitDate_lable = Label(root, text="Book submit Date", font=font1).place(x=400, y=350)
    bookSubmitDate_Entry = Entry(root, textvariable=bookSubmitDate_var, font=font1).place(x=550, y=350)

    bookLateSubmitDate_lable = Label(root, text="Book late Submit Date", font=font1).place(x=400, y=400)
    bookLateSubmitDate_Entry = Entry(root, textvariable=bookLateSubmitDate_var, font=font1).place(x=550, y=400)

    booklateSubmmitFine_lable = Label(root, text="Late Submit fine", font=font1).place(x=400, y=450)
    booklateSubmmitFine_Entry = Entry(root, textvariable=booklateSubmmitFine_var, font=font1).place(x=550, y=450)

    totalPricepad_lable = Label(root, text="Total Price", font=font1).place(x=400, y=500)
    totalPricepad_Entry = Entry(root, textvariable=totalPricepad_var, font=font1).place(x=550, y=500)

    sub_btn = Button(root, text='SUBMIT', width=16, font=font1, bg="#02de53"  , command=submit)
    sub_btn.place(x=300, y=550)

    root.mainloop()
